import io
import csv
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, update
from datetime import datetime, timezone
from app.models.references import DictionaryEntry, BrandWhitelist, ChecklistRule, RuleCategory
from app.core.database import get_db
from app.core.deps import require_admin
from app.models.users import User
from app.models.audit_log import AuditLog
from app.models.checks import CheckTask, TaskStatus
from app.services import config_service
from app.pipeline.providers import get_ocr_provider, get_llm_provider
from app.services.storage import StorageService, get_storage_service

router = APIRouter(prefix="/admin", tags=["admin"])


@router.get("/config")
async def get_config(db: AsyncSession = Depends(get_db), _: User = Depends(require_admin)):
    api_keys = await config_service.get_masked_api_keys(db)
    pipeline = await config_service.get_pipeline_config(db)
    s3_endpoint = await config_service.get_config(db, "s3_endpoint_url") or ""
    s3_bucket = await config_service.get_config(db, "s3_bucket") or ""
    extras = await config_service.get_extras(db)
    debug_mode = (await config_service.get_config(db, "debug_mode")) == "true"
    providers_available = {}
    for p in config_service.API_KEY_PROVIDERS:
        providers_available[p] = bool(await config_service.get_config(db, f"api_key_{p}"))
    return {"api_keys": api_keys, "pipeline": pipeline,
            "s3": {"endpoint_url": s3_endpoint, "bucket": s3_bucket}, "extras": extras,
            "debug_mode": debug_mode, "providers_available": providers_available}


@router.put("/config")
async def update_config(
    payload: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    api_keys = payload.get("api_keys", {})
    for provider, key_value in api_keys.items():
        if key_value and not key_value.startswith("****"):
            await config_service.set_config(db, f"api_key_{provider}", key_value, is_encrypted=True, updated_by_id=current_user.id)

    pipeline = payload.get("pipeline", {})
    for k, v in pipeline.items():
        await config_service.set_config(db, k, v or "", updated_by_id=current_user.id)

    s3 = payload.get("s3", {})
    for k, v in s3.items():
        encrypted = k in ("access_key", "secret_key")
        await config_service.set_config(db, f"s3_{k}", v or "", is_encrypted=encrypted, updated_by_id=current_user.id)

    if "debug_mode" in payload:
        await config_service.set_config(db, "debug_mode", "true" if payload["debug_mode"] else "false", updated_by_id=current_user.id)

    extras = payload.get("extras", {})
    if "yandex_folder_id" in extras:
        await config_service.set_config(db, "yandex_folder_id", extras.get("yandex_folder_id") or "", updated_by_id=current_user.id)
    if "abbyy_url" in extras and extras.get("abbyy_url"):
        await config_service.set_config(db, "abbyy_url", extras["abbyy_url"], updated_by_id=current_user.id)
    if extras.get("abbyy_password") and not extras["abbyy_password"].startswith("****"):
        await config_service.set_config(db, "abbyy_password", extras["abbyy_password"], is_encrypted=True, updated_by_id=current_user.id)

    return {"message": "Конфигурация сохранена"}


@router.post("/config/test-connection")
async def test_connection(
    payload: dict,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_admin),
):
    provider_type = payload.get("provider_type", "llm")
    provider_name = payload.get("provider_name", "")
    try:
        if provider_type == "storage":
            ok = (await get_storage_service(db)).test_connection()
        elif provider_type == "ocr":
            ok = await (await config_service.build_ocr_provider(db, provider_name)).test_connection()
        else:
            ok = await (await config_service.build_llm_provider(db, provider_name)).test_connection()
        return {"success": ok, "message": "Подключение успешно" if ok else "Ошибка подключения"}
    except Exception as e:
        return {"success": False, "message": str(e)}


@router.post("/purge-queue")
async def purge_queue(db: AsyncSession = Depends(get_db), _: User = Depends(require_admin)):
    """Clear the broker queue and mark stale PENDING checks as FAILED (drains zombie backlog)."""
    from app.workers.celery_app import celery_app
    purged = None
    try:
        with celery_app.connection_for_write() as conn:
            purged = conn.default_channel.queue_purge("celery")
    except Exception as e:
        purged = f"err: {e}"
    res = await db.execute(update(CheckTask).where(CheckTask.status == TaskStatus.PENDING).values(
        status=TaskStatus.FAILED, error="stale/purged", completed_at=datetime.now(timezone.utc)))
    await db.commit()
    return {"purged_messages": purged, "pending_marked_failed": res.rowcount}


def _parse_rows(filename: str, content: bytes) -> list[list[str]]:
    """Return rows (list of cell-string lists) from .xlsx or .csv content."""
    if filename.lower().endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        return [[("" if c is None else str(c)).strip() for c in row] for row in ws.iter_rows(values_only=True)]
    text = content.decode("utf-8-sig", errors="replace")
    delim = ";" if text.count(";") > text.count(",") else ","
    return [[c.strip() for c in row] for row in csv.reader(io.StringIO(text), delimiter=delim)]


@router.post("/import/{kind}")
async def import_reference(
    kind: str,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    """Import dictionary terms / brands / checklist rules from .xlsx or .csv (skips duplicates)."""
    if kind not in ("dictionary", "brands", "checklist"):
        raise HTTPException(400, "kind: dictionary | brands | checklist")
    rows = _parse_rows(file.filename or "f.csv", await file.read())
    # Drop a header row if the first cell looks like a header.
    if rows and rows[0] and rows[0][0].lower() in ("term", "термин", "слово", "brand", "бренд", "rule_key", "ключ"):
        rows = rows[1:]
    added = 0
    for r in rows:
        if not r or not r[0]:
            continue
        try:
            if kind == "dictionary":
                term = r[0]
                if not (await db.execute(select(DictionaryEntry).where(DictionaryEntry.term == term))).scalar_one_or_none():
                    db.add(DictionaryEntry(term=term, category=(r[1] if len(r) > 1 and r[1] else "general"))); added += 1
            elif kind == "brands":
                if not (await db.execute(select(BrandWhitelist).where(BrandWhitelist.brand_name == r[0]))).scalar_one_or_none():
                    db.add(BrandWhitelist(brand_name=r[0])); added += 1
            else:  # checklist: rule_key, description, [category]
                key = r[0]
                desc = r[1] if len(r) > 1 else r[0]
                cat = (r[2].lower() if len(r) > 2 and r[2] else "all")
                cat = cat if cat in RuleCategory._value2member_map_ else "all"
                if not (await db.execute(select(ChecklistRule).where(ChecklistRule.rule_key == key))).scalar_one_or_none():
                    db.add(ChecklistRule(rule_key=key, description=desc, category=RuleCategory(cat))); added += 1
        except Exception:
            continue
    await db.commit()
    return {"kind": kind, "rows": len(rows), "added": added}


@router.get("/logs")
async def get_logs(
    skip: int = 0,
    limit: int = 100,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(require_admin),
):
    result = await db.execute(select(AuditLog).order_by(AuditLog.created_at.desc()).offset(skip).limit(limit))
    logs = result.scalars().all()
    return [{"id": str(l.id), "user_id": str(l.user_id) if l.user_id else None, "action": l.action, "resource_type": l.resource_type, "resource_id": l.resource_id, "ip": l.ip_address, "created_at": l.created_at.isoformat()} for l in logs]


@router.get("/stats")
async def get_stats(db: AsyncSession = Depends(get_db), _: User = Depends(require_admin)):
    total = (await db.execute(select(func.count(CheckTask.id)))).scalar_one()
    completed = (await db.execute(select(func.count(CheckTask.id)).where(CheckTask.status == TaskStatus.COMPLETED))).scalar_one()
    failed = (await db.execute(select(func.count(CheckTask.id)).where(CheckTask.status == TaskStatus.FAILED))).scalar_one()
    return {"total_checks": total, "completed": completed, "failed": failed, "pending": total - completed - failed}
